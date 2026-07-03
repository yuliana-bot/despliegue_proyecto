"""
UTILIDADES PARA EXPORTACION DE REPORTES
"""
import os
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
from datetime import datetime

# ══════════════════════════════════════════════════════════
#  RUTAS DE LOGO
# ══════════════════════════════════════════════════════════

LOGO_PATH = 'file:///' + os.path.join(
    settings.BASE_DIR, 'app', 'static', 'imagenes', 'logoempresa.jpeg'
).replace('\\', '/')

LOGO_PATH_EXCEL = os.path.join(
    settings.BASE_DIR, 'app', 'static', 'imagenes', 'logoempresa.jpeg'
)

# ══════════════════════════════════════════════════════════
#  COLORES — igual que el PDF
# ══════════════════════════════════════════════════════════
C_NEGRO       = '1A1A1A'   # Header lateral y encabezado tabla
C_ROJO        = 'D32F2F'   # Header central y acento
C_ROJO_OSCURO = 'B71C1C'   # Borde inferior header / separador
C_BLANCO      = 'FFFFFF'
C_GRIS_CLARO  = 'F2F2F0'   # Filas impares / barra resumen
C_GRIS_BORDE  = 'E0DDD8'   # Bordes tabla
C_TEXTO       = '2A2A2A'   # Texto datos


def _side(color, style='thin'):
    return Side(style=style, color=color)

def _border(color='E0DDD8', style='thin'):
    s = _side(color, style)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def _font(size=10, bold=False, color='2A2A2A', name='Arial'):
    return Font(name=name, size=size, bold=bold, color=color)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ══════════════════════════════════════════════════════════
#  EXPORTAR PDF
# ══════════════════════════════════════════════════════════

def exportar_pdf(titulo, columnas, datos, nombre_archivo):
    contexto = {
        'titulo': titulo,
        'columnas': columnas,
        'datos': datos,
        'logo_url': LOGO_PATH,
    }
    html_string = render_to_string('reportes/reporte_pdf.html', contexto)
    html_object = HTML(string=html_string, base_url='.')
    pdf_bytes = html_object.write_pdf()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    return response


# ══════════════════════════════════════════════════════════
#  EXPORTAR EXCEL  — diseño igual al PDF nuevo
# ══════════════════════════════════════════════════════════

def exportar_excel(titulo, columnas, datos, nombre_archivo, generado_por=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    num_cols   = len(columnas)
    ultima_col = get_column_letter(num_cols)
    fecha_txt  = datetime.now().strftime('%d/%m/%Y')
    hora_txt   = datetime.now().strftime('%H:%M')
    gen_txt    = str(generado_por) if generado_por else 'Sistema'

    # ── Ocultar líneas de cuadrícula ──────────────────────
    ws.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════
    #  FILA 1-3 — HEADER en 3 bloques: NEGRO | ROJO | NEGRO
    #  (igual que el PDF: logo | ACERAUTOS | fecha/hora)
    # ══════════════════════════════════════════════════════
    HEADER_ROWS = 3
    for r in range(1, HEADER_ROWS + 1):
        ws.row_dimensions[r].height = 22

    # Ancho proporcional de columnas del header
    col_logo_fin  = max(1, num_cols // 5)            # ~20% para logo
    col_rojo_ini  = col_logo_fin + 1
    col_rojo_fin  = num_cols - max(1, num_cols // 5) # ~60% para nombre empresa
    col_fecha_ini = col_rojo_fin + 1                 # ~20% para fecha

    # Bloque NEGRO izquierdo (logo)
    if col_logo_fin >= 1:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=HEADER_ROWS, end_column=col_logo_fin)
        c = ws.cell(row=1, column=1)
        c.fill      = _fill(C_NEGRO)
        c.alignment = _align('center', 'center')

        if os.path.exists(LOGO_PATH_EXCEL):
            try:
                img        = XLImage(LOGO_PATH_EXCEL)
                img.height = 58
                img.width  = 100
                img.anchor = 'A1'
                ws.add_image(img)
            except Exception:
                c.value = 'ACERAUTOS'
                c.font  = _font(11, True, C_BLANCO)

    # Bloque ROJO central (nombre empresa)
    if col_rojo_fin >= col_rojo_ini:
        ws.merge_cells(start_row=1, start_column=col_rojo_ini,
                       end_row=HEADER_ROWS, end_column=col_rojo_fin)
        c             = ws.cell(row=1, column=col_rojo_ini)
        c.value       = 'ACERAUTOS'
        c.font        = Font(name='Arial', size=22, bold=True, color=C_BLANCO)
        c.fill        = _fill(C_ROJO)
        c.alignment   = _align('center', 'center')

    # Bloque NEGRO derecho (fecha / hora)
    if col_fecha_ini <= num_cols:
        ws.merge_cells(start_row=1, start_column=col_fecha_ini,
                       end_row=HEADER_ROWS, end_column=num_cols)
        c             = ws.cell(row=1, column=col_fecha_ini)
        c.value       = f'Fecha:\n{fecha_txt}\n\nHora:\n{hora_txt}'
        c.font        = Font(name='Arial', size=9, color=C_BLANCO)
        c.fill        = _fill(C_NEGRO)
        c.alignment   = Alignment(horizontal='right', vertical='center',
                                  wrap_text=True)

    # ── Línea separadora roja debajo del header ───────────
    FILA_SEP = HEADER_ROWS + 1
    ws.row_dimensions[FILA_SEP].height = 4
    ws.merge_cells(start_row=FILA_SEP, start_column=1,
                   end_row=FILA_SEP, end_column=num_cols)
    ws.cell(row=FILA_SEP, column=1).fill = _fill(C_ROJO_OSCURO)

    # ══════════════════════════════════════════════════════
    #  FILA TÍTULO DEL REPORTE
    # ══════════════════════════════════════════════════════
    FILA_TITULO = FILA_SEP + 1
    ws.row_dimensions[FILA_TITULO].height = 26
    ws.merge_cells(start_row=FILA_TITULO, start_column=1,
                   end_row=FILA_TITULO, end_column=num_cols)
    c           = ws.cell(row=FILA_TITULO, column=1)
    c.value     = titulo.upper()
    c.font      = Font(name='Arial', size=13, bold=True, color=C_BLANCO)
    c.fill      = _fill(C_ROJO)
    c.alignment = _align('center', 'center')

    # ── Franja gris con "Generado por" ────────────────────
    FILA_META = FILA_TITULO + 1
    ws.row_dimensions[FILA_META].height = 15
    ws.merge_cells(start_row=FILA_META, start_column=1,
                   end_row=FILA_META, end_column=num_cols)
    c           = ws.cell(row=FILA_META, column=1)
    c.value     = f'Generado por: {gen_txt}'
    c.font      = Font(name='Arial', size=8, italic=True, color='888888')
    c.fill      = _fill(C_GRIS_CLARO)
    c.alignment = _align('right', 'center')

    # ══════════════════════════════════════════════════════
    #  ENCABEZADOS DE TABLA — fondo NEGRO texto blanco
    # ══════════════════════════════════════════════════════
    FILA_ENC = FILA_META + 1
    ws.row_dimensions[FILA_ENC].height = 20

    borde_enc_bot = Border(
        bottom=Side(style='medium', color=C_ROJO),
        left=_side(C_GRIS_BORDE),
        right=_side(C_GRIS_BORDE),
        top=_side(C_NEGRO),
    )

    for col_i, col_name in enumerate(columnas, 1):
        c           = ws.cell(row=FILA_ENC, column=col_i)
        c.value     = str(col_name).upper()
        c.font      = Font(name='Arial', size=9, bold=True, color=C_BLANCO)
        c.fill      = _fill(C_NEGRO)
        c.alignment = _align('center', 'center')
        c.border    = borde_enc_bot
        # Acento rojo en primera columna (como el PDF)
        if col_i == 1:
            c.border = Border(
                left=Side(style='thick', color=C_ROJO),
                right=_side(C_GRIS_BORDE),
                top=_side(C_NEGRO),
                bottom=Side(style='medium', color=C_ROJO),
            )

    # ══════════════════════════════════════════════════════
    #  FILAS DE DATOS — alternas gris claro / blanco
    # ══════════════════════════════════════════════════════
    FILA_DATOS_INI = FILA_ENC + 1

    for idx, fila in enumerate(datos):
        row_num = FILA_DATOS_INI + idx
        ws.row_dimensions[row_num].height = 16
        color_fila = C_GRIS_CLARO if idx % 2 == 0 else C_BLANCO
        valores = list(fila) if not isinstance(fila, dict) else \
                  [fila.get(col.lower().replace(' ', '_'), '') for col in columnas]

        for col_i, valor in enumerate(valores, 1):
            c           = ws.cell(row=row_num, column=col_i)
            c.value     = valor
            c.font      = _font(10, False, C_TEXTO)
            c.fill      = _fill(color_fila)
            c.alignment = _align('left', 'center')
            c.border    = Border(
                bottom=_side(C_GRIS_BORDE),
                left=_side(C_GRIS_BORDE),
                right=_side(C_GRIS_BORDE),
            )
            # Acento rojo en primera columna
            if col_i == 1:
                c.border = Border(
                    left=Side(style='thick', color=C_ROJO),
                    right=_side(C_GRIS_BORDE),
                    bottom=_side(C_GRIS_BORDE),
                )

    # ══════════════════════════════════════════════════════
    #  BARRA DE RESUMEN — igual que el PDF
    # ══════════════════════════════════════════════════════
    FILA_RES = FILA_DATOS_INI + len(datos) + 1
    ws.row_dimensions[FILA_RES].height = 20

    borde_res = Border(
        top=Side(style='medium', color=C_ROJO),
        bottom=_side(C_GRIS_BORDE),
        left=_side(C_GRIS_BORDE),
        right=_side(C_GRIS_BORDE),
    )

    t1 = max(num_cols // 3, 1)
    t2 = t1 + 1
    t3 = max((num_cols * 2) // 3, t2)
    t4 = min(t3 + 1, num_cols)

    # Total registros
    ws.merge_cells(start_row=FILA_RES, start_column=1,
                   end_row=FILA_RES, end_column=t1)
    c           = ws.cell(row=FILA_RES, column=1)
    c.value     = f'Total de registros: {len(datos)}'
    c.font      = Font(name='Arial', size=9, bold=True, color=C_ROJO)
    c.fill      = _fill(C_GRIS_CLARO)
    c.alignment = _align('left', 'center')
    c.border    = Border(top=Side(style='medium', color=C_ROJO),
                         left=Side(style='thick', color=C_ROJO))

    # Nombre empresa
    if t3 >= t2:
        ws.merge_cells(start_row=FILA_RES, start_column=t2,
                       end_row=FILA_RES, end_column=t3)
    c           = ws.cell(row=FILA_RES, column=t2)
    c.value     = 'Acerautos — Centro Integral Automotriz'
    c.font      = Font(name='Arial', size=9, bold=True, color=C_TEXTO)
    c.fill      = _fill(C_GRIS_CLARO)
    c.alignment = _align('center', 'center')
    c.border    = borde_res

    # Fecha generación
    if t4 <= num_cols:
        ws.merge_cells(start_row=FILA_RES, start_column=t4,
                       end_row=FILA_RES, end_column=num_cols)
    c           = ws.cell(row=FILA_RES, column=t4)
    c.value     = f'Generado el: {fecha_txt}  {hora_txt}'
    c.font      = Font(name='Arial', size=9, color='555555')
    c.fill      = _fill(C_GRIS_CLARO)
    c.alignment = _align('right', 'center')
    c.border    = borde_res

    # ══════════════════════════════════════════════════════
    #  AJUSTE AUTOMÁTICO DE ANCHOS
    # ══════════════════════════════════════════════════════
    for col_i, col_name in enumerate(columnas, 1):
        col_letra  = get_column_letter(col_i)
        max_length = len(str(col_name))
        for row in ws.iter_rows(min_row=FILA_ENC, max_row=FILA_DATOS_INI + len(datos),
                                min_col=col_i, max_col=col_i):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letra].width = min(max_length + 4, 45)

    # ── Respuesta HTTP ────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)
    return response